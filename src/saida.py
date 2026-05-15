from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

PASTA_SAIDA = Path("saida")
CABECALHOS = [
    "Codigo Pessoa", "Nome Entrada",
    "Situacao", "Filial", "Codigo", "Pessoa", "Nome Systur", "CPF", "Telefone"
]
COR_HEADER = "D3D3D3"  # cinza claro


def criar_arquivo() -> tuple[openpyxl.Workbook, object, Path]:
    PASTA_SAIDA.mkdir(exist_ok=True)
    agora = datetime.now().strftime("%d_%m_%H_%M_%S")
    caminho = PASTA_SAIDA / f"vendedores_{agora}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    font = Font(color="000000", bold=True)

    for col, titulo in enumerate(CABECALHOS, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    wb.save(caminho)
    print(f"[SAIDA] Arquivo criado: {caminho.name}")
    return wb, ws, caminho


def adicionar_linhas(ws, codigo_pessoa: int, nome_entrada: str, resultados: list[dict]) -> None:
    for r in resultados:
        ws.append([
            codigo_pessoa,
            nome_entrada,
            r.get("situacao", ""),
            r.get("filial", ""),
            r.get("codigo", ""),
            r.get("pessoa", ""),
            r.get("nome", ""),
            r.get("cpf", ""),
            r.get("telefone", ""),
        ])


def salvar(wb: openpyxl.Workbook, caminho: Path) -> None:
    wb.save(caminho)
